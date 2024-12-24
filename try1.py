import requests
import pandas as pd
from openpyxl import load_workbook

# API URL and Token
API_URL = "https://trefle.io/api/v1/plants"
API_TOKEN = "8tgKi4Kj1fMQOwvBOUgSsLv2fwfq0xpsi47h_K8Sv7c"
params = {
    "token": API_TOKEN,
    "filter[edible]": "true"
}

# Fetch Data from Multiple Pages
plants = []
for page in range(1, 11):  # Loop through 10 pages
    params["page"] = page
    response = requests.get(API_URL, params=params)
    if response.status_code == 200:
        data = response.json().get("data", [])
        for plant in data:
            plants.append({
                "ID": plant.get("id"),
                "Common Name": plant.get("common_name"),
                "Scientific Name": plant.get("scientific_name"),
                "Year": plant.get("year"),
                "Author": plant.get("author"),
                "Status": plant.get("status"),
                "Rank": plant.get("rank"),
                "Family Common Name": plant.get("family_common_name"),
                "Genus": plant.get("genus"),
                "Image URL": plant.get("image_url"),
                "Edible": True  # Adding extra field
            })
    else:
        print(f"Failed to fetch page {page}: {response.status_code}")

# Create DataFrame
new_data = pd.DataFrame(plants)

# File to Append Data
output_file = "edible_plants1.xlsx"

try:
    # Load existing workbook and determine the last row
    existing_workbook = load_workbook(output_file)
    sheet_name = 'Sheet1'

    if sheet_name in existing_workbook.sheetnames:
        startrow = existing_workbook[sheet_name].max_row
    else:
        startrow = 0

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # Append new data
        new_data.to_excel(writer, index=False, header=False if startrow > 0 else True, startrow=startrow, sheet_name=sheet_name)

    print(f"Data appended to {output_file}")

except FileNotFoundError:
    # File does not exist, create a new one
    new_data.to_excel(output_file, index=False)
    print(f"File {output_file} created and data written.")
